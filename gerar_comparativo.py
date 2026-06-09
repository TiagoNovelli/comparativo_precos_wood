#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera o relatorio HTML "Comparativo de Precos Wood" a partir das planilhas:
  - Novos precos.xlsx                                    (precos novos, fonte unica)
  - 1 - TABELA DE PRECOS WOOD - CENTURY 28.04.xlsx       (precos antigos, marca Century)
  - 1 - TABELA DE PRECOS WOOD - PV 28.04.xlsx            (precos antigos, marca Pv)
  - 1 - TABELA DE PRECOS WOOD - PRIVATE LABEL 28.04.xlsx (precos antigos, marca Private Label)
  - Precos antigos - Century e .pv.xlsx                  (fonte legada, usada como FALLBACK)

Regra de correspondencia (fixa, deterministica), em duas etapas:
  1) Planilhas de marca: para cada produto novo (MODELO ... + categoria), procura
     nas planilhas antigas da MESMA categoria um produto cujo nome compartilhe a
     "palavra-chave" do modelo; entre todos os pares (variante nova x produto
     antigo candidato), escolhe o par com a MENOR diferenca de dimensoes (C+L+A).
     Os acabamentos novos sao mapeados para as colunas fixas de acabamento dessas
     planilhas (FINISH_MAP / ESP_FINISH_MAP), com as correcoes de Marmore Fornecido
     (-8%) e Vidro Normal (= preco "sem topo").
  2) Fonte legada (somente se a etapa 1 nao encontrar correspondencia): mesma
     regra de nome-chave + menor diferenca dimensional, mas SEM filtro de
     categoria (a planilha legada nao separa por categoria/aba e inclui
     acessorios como ESPELHO/MANCEBO). Ali o tipo de acabamento do tampo vem
     descrito dentro do texto da "configuracao" (ex.: "TIPO DE ACABAMENTO DO
     TAMPO: LAMINADO"), no mesmo vocabulario das colunas de "Novos precos" -
     entao a comparacao e feita por nome direto (ver LEGACY_FINISH_ALIASES para
     pequenas variacoes de grafia). Quando o produto novo nao diferencia
     acabamentos (so tem o preco "Nenhum" preenchido - ex.: modelos da linha
     SAMMY), esse preco unico e comparado contra CADA acabamento disponivel na
     configuracao antiga encontrada. Quando um acabamento existe de um lado mas
     nao do outro, a linha aparece como informativa (sem calculo de %).
  Sem candidato em nenhuma das fontes -> o produto cai em "Sem Correspondente",
  onde e exibida apenas a configuracao de tamanho mais representativa do produto
  (variante mais proxima da dimensao media do produto, ou de 2,70m de comprimento
  no caso de mesas de jantar) com os precos novos, sem comparacao.

Marmores Especiais (ESP): o relatorio e unico, com um interruptor no topo que
  alterna - via JS/CSS, sem nova geracao - entre incluir ou excluir os acabamentos
  ESP da analise; medias, badges e contadores sao recalculados (pre-computados
  nas duas variantes e trocados pelo interruptor).

Limitacoes conhecidas (vs. o relatorio anterior, que teve curadoria manual):
  - Nao reproduz a logica combinatoria especial da "Mesa Eclipse" (tampo
    inferior+superior); ela e tratada como produto comum.
  - As combinacoes podem diferir do relatorio anterior em casos ambiguos, pois
    aquele envolveu escolhas manuais (ex.: variantes de tamanho escolhidas a
    dedo, referencias cross-categoria).

Uso:
    python gerar_comparativo.py
Gera (sobrescrevendo) "Comparativo Preços Wood.html" na mesma pasta.
"""
import re
import unicodedata
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore", message="Data Validation extension is not supported")

BASE = Path(__file__).resolve().parent

NEW_PRICES_FILE = BASE / "Novos preços.xlsx"
OLD_BRAND_FILES = {
    "Century": BASE / "1 - TABELA DE PRECOS WOOD - CENTURY 28.04.xlsx",
    "Pv": BASE / "1 - TABELA DE PRECOS WOOD - PV 28.04.xlsx",
    "Private Label": BASE / "1 - TABELA DE PRECOS WOOD - PRIVATE LABEL 28.04.xlsx",
}
# fonte legada (fallback): usada apenas quando o produto nao casa com nenhuma das
# 3 planilhas de marca. Estrutura propria (modelo/modulacao/configuracao/preco) -
# o tipo de acabamento do tampo vem descrito dentro do texto da "configuracao",
# nao em colunas fixas como nas planilhas de marca.
LEGACY_PRICES_FILE = BASE / "Preços antigos - Century e .pv.xlsx"

OUTPUT_REPORT = BASE / "Comparativo Preços Wood.html"

# categoria (nome usado no relatorio) -> nome da aba nas planilhas antigas
CATEGORY_SHEETS = {
    "Mesa de Jantar": "MESA DE JANTAR",
    "Mesa de Centro": "MESA DE CENTRO",
    "Mesa de Cabeceira": "MESA DE CABECEIRA",
    "Mesa Lateral": "MESA LATERAL",
    "Aparador": "APARADOR",
    "Buffet": "BUFFET",
    "Cômoda": "CÔMODA",
}
CATEGORY_ORDER = list(CATEGORY_SHEETS)

# "MODULACAO DO PRODUTO" (planilha de precos novos) -> categoria do relatorio
MODULACAO_TO_CATEGORY = {
    "MESA JANTAR": "Mesa de Jantar",
    "MESA CENTRO": "Mesa de Centro",
    "MESA CABECEIRA": "Mesa de Cabeceira",
    "MESA LATERAL": "Mesa Lateral",
    "APARADOR": "Aparador",
    "BUFFET": "Buffet",
    "COMODA": "Cômoda",
}

# indices fixos das colunas de acabamento nas planilhas antigas (linha de cabecalho
# = "Imagem | Nome | C | L | A | (vazia) | <5 colunas de tampo> ")
COL_VIDRO_ESPELHO = 6     # Tampo: Vidro/Espelho ...
COL_LACA_VIDRO_NORMAL = 7  # Tampo: Laca/Lamina com ou sem Vidro Normal
COL_MARMORE_ESPECIAL = 8   # Tampo: Marmore (ou Porcelana) Especial
COL_MARMORE_NORMAL = 9     # Tampo: Marmore (ou Porcelana) Normal
HEADER_ROW = 6
DATA_START_ROW = 7

MARMORE_FORNECIDO_FACTOR = 0.92  # correcao: preco "sem topo" -8%

# acabamento novo -> (coluna antiga correspondente, fator de correcao)
FINISH_MAP = {
    "NENHUM": (COL_LACA_VIDRO_NORMAL, 1.0),
    "VIDRO": (COL_LACA_VIDRO_NORMAL, 1.0),
    "ESPELHO": (COL_VIDRO_ESPELHO, 1.0),
    "VIDRO FOSCO": (COL_VIDRO_ESPELHO, 1.0),
    "VIDRO EXTRACLEAR": (COL_VIDRO_ESPELHO, 1.0),
    "MARMORE POLIDO": (COL_MARMORE_NORMAL, 1.0),
    "MARMORE LEVIGADO": (COL_MARMORE_NORMAL, 1.0),
    "MARMORE ESCOVADO": (COL_MARMORE_NORMAL, 1.0),
    "MARMORE FORNECIDO": (COL_LACA_VIDRO_NORMAL, MARMORE_FORNECIDO_FACTOR),
}
ESP_FINISH_MAP = {
    "MARMORE POLIDO ESP": (COL_MARMORE_ESPECIAL, 1.0),
    "MARMORE LEVIGADO ESP": (COL_MARMORE_ESPECIAL, 1.0),
    "MARMORE ESCOVADO ESP": (COL_MARMORE_ESPECIAL, 1.0),
}

# Equivalencias entre o nome do acabamento em "Novos precos" (colunas) e a forma
# como aparece descrito na "configuracao" da planilha legada "Precos antigos -
# Century e .pv" (mesmo vocabulario, com pequenas variacoes de grafia).
LEGACY_FINISH_ALIASES = {
    "VIDRO EXTRACLEAR 4MM": "VIDRO EXTRACLEAR",
}

# ordem de exibicao das linhas de acabamento dentro de cada bloco de produto
# (sempre inclui ESP - a visibilidade/calculo "com ou sem ESP" e controlada pelo
# interruptor no topo do relatorio, client-side)
FINISH_ORDER_ESP = [
    "NENHUM", "VIDRO", "ESPELHO", "VIDRO FOSCO", "VIDRO EXTRACLEAR",
    "MARMORE POLIDO", "MARMORE LEVIGADO", "MARMORE ESCOVADO",
    "MARMORE POLIDO ESP", "MARMORE LEVIGADO ESP", "MARMORE ESCOVADO ESP",
    "MARMORE FORNECIDO",
]

# observacao exibida ao lado de certos acabamentos
FINISH_NOTES = {
    "VIDRO": "vidro normal = sem topo (base antiga)",
    "MARMORE FORNECIDO": "sem topo −8%",
}

NEUTRAL_THRESHOLD = 2.0  # |variacao %| < 2.0 => "neutral"

# palavras genericas ignoradas ao identificar o "nome-chave" de um produto
STOPWORDS = {
    "MESA", "DE", "DA", "DO", "E", "OU", "COM", "SEM", "ACABAMENTOS", "ACABAMENTO",
    "CUSTOMIZADOS", "CUSTOMIZADO", "JANTAR", "CENTRO", "CABECEIRA", "LATERAL",
    "BUFFET", "APARADOR", "COMODA", "RETANGULAR", "ORGANICA", "REDONDO",
    "MODELO", "TOTAL", "BAR",
}


# --------------------------------------------------------------------------- #
# utilidades
# --------------------------------------------------------------------------- #

def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def core_tokens(name):
    """Tokens significativos (maiusculos, sem acento/pontuacao, sem stopwords)."""
    cleaned = strip_accents(name).upper()
    cleaned = re.sub(r"[^A-Z0-9 ]", " ", cleaned)
    tokens = [t for t in cleaned.split() if t and t not in STOPWORDS]
    return tokens


def fmt_brl(value):
    s = f"{value:,.2f}"
    s = s.replace(",", "§").replace(".", ",").replace("§", ".")
    return f"R$ {s}"


def fmt_brl_signed(value):
    sign = "+" if value >= 0 else "-"
    return f"{sign}{fmt_brl(abs(value))}"


def fmt_pct(value):
    sign = "+" if value >= 0 else ""
    return f"{sign}{value:.1f}%"


def variation_class(pct):
    if abs(pct) < NEUTRAL_THRESHOLD:
        return "neutral"
    return "up" if pct > 0 else "down"


def arrow(pct):
    return "↑" if pct >= 0 else "↓"


def fmt_size(c, l, a):
    a_str = f"{a:.2f}" if a is not None else ""
    c_str = f"{c:.2f}" if c is not None else ""
    l_str = f"{l:.2f}" if l is not None else ""
    return f"{c_str}x{l_str}x{a_str}m"


def short_name(display_name):
    if display_name.startswith("Mesa "):
        return display_name[len("Mesa "):].upper()
    return display_name.upper()


def category_label(categoria_raw):
    """Nome de categoria para exibicao: usa o mapeamento conhecido; para
    modulacoes fora das 7 categorias principais (ex.: ESPELHO, MANCEBO,
    SOFA TABLE - normalmente acessorios encontrados via fonte legada), deriva
    um titulo a partir do texto bruto da modulacao."""
    if categoria_raw in MODULACAO_TO_CATEGORY:
        return MODULACAO_TO_CATEGORY[categoria_raw]
    return (categoria_raw or "Outros").title()


def cat_slug(cat):
    return strip_accents(cat).lower().replace(" ", "-")


def normalize_finish_name(name):
    return strip_accents(name).upper().strip()


# --------------------------------------------------------------------------- #
# leitura: precos novos
# --------------------------------------------------------------------------- #

CONFIG_FIELDS = {
    "MODULACAO DO PRODUTO": "categoria_raw",
    "FORMATO DA MODULAÇÃO": "formato",
    "COMPRIMENTO DA MODULACAO": "C",
    "PROFUNDIDADE DA MODULACAO": "L",
    "ALTURA DA MODULACAO": "A",
}


def parse_config(text):
    info = {"categoria_raw": None, "formato": None, "C": None, "L": None, "A": None}
    if not text:
        return info
    cleaned = strip_accents(text).upper()
    for line in cleaned.splitlines():
        line = line.strip()
        if ":" not in line:
            continue
        key, _, val = line.partition(":")
        key, val = key.strip(), val.strip()
        if key == "MODULACAO DO PRODUTO":
            info["categoria_raw"] = val
        elif key == "FORMATO DA MODULACAO":
            info["formato"] = val
        elif key in ("COMPRIMENTO DA MODULACAO", "PROFUNDIDADE DA MODULACAO", "ALTURA DA MODULACAO"):
            m = re.search(r"[\d.]+", val)
            if m:
                num = float(m.group(0))
                if key.startswith("COMPRIMENTO"):
                    info["C"] = num
                elif key.startswith("PROFUNDIDADE"):
                    info["L"] = num
                else:
                    info["A"] = num
    return info


def load_new_products():
    """Retorna lista de produtos novos: cada um com nome de exibicao e variantes."""
    wb = openpyxl.load_workbook(NEW_PRICES_FILE, data_only=True)
    ws = wb["Export"]
    header = [c.value for c in ws[1]]

    groups = {}   # modelo_raw -> {"display": ..., "variants": [...]}
    order = []
    current = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        modelo_raw, config_text = row[0], row[1]
        if modelo_raw:
            current = modelo_raw
        if current is None:
            continue
        if current not in groups:
            display = current.replace("MODELO ", "", 1).title()
            groups[current] = {"display": display, "variants": []}
            order.append(current)

        info = parse_config(config_text)
        finishes = {}
        for name, value in zip(header[2:], row[2:]):
            if name and value is not None:
                finishes[name] = float(value)
        groups[current]["variants"].append({
            "categoria_raw": info["categoria_raw"],
            "formato": info["formato"],
            "C": info["C"], "L": info["L"], "A": info["A"],
            "finishes": finishes,
        })

    products = []
    for key in order:
        g = groups[key]
        # agrupa variantes por categoria (um mesmo modelo pode existir em > 1 categoria)
        by_cat = {}
        cat_order = []
        for v in g["variants"]:
            raw = v["categoria_raw"] or "?"
            if raw not in by_cat:
                by_cat[raw] = []
                cat_order.append(raw)
            by_cat[raw].append(v)
        for raw in cat_order:
            products.append({
                "display": g["display"],
                "categoria_raw": raw,
                "categoria": MODULACAO_TO_CATEGORY.get(raw),
                "variants": by_cat[raw],
            })
    return products


# --------------------------------------------------------------------------- #
# leitura: precos antigos (3 planilhas de marca)
# --------------------------------------------------------------------------- #

def load_old_products():
    """Retorna dict: categoria -> lista de produtos antigos
    {brand, name, C, L, A, cols: [v6..v10]}"""
    by_category = {cat: [] for cat in CATEGORY_SHEETS}
    for brand, path in OLD_BRAND_FILES.items():
        wb = openpyxl.load_workbook(path, data_only=True)
        for cat, sheet_name in CATEGORY_SHEETS.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row, values_only=True):
                name = row[1]
                if not name or not isinstance(name, str):
                    continue
                c, l, a = row[2], row[3], row[4]
                cols = [row[6], row[7], row[8], row[9], row[10]] if len(row) > 10 else (list(row[6:10]) + [None])
                if all(v is None for v in cols):
                    continue
                by_category[cat].append({
                    "brand": brand, "name": name.strip(),
                    "C": c, "L": l, "A": a, "cols": cols,
                })
    return by_category


def numeric_dim(value, fallback):
    """Converte dimensao para numero; usa 'fallback' quando o valor nao e numerico
    (ex.: o simbolo de diametro em mesas redondas, onde C='diametro' e L=valor)."""
    if isinstance(value, (int, float)):
        return float(value)
    return fallback


def dimension_diff(new_v, old_p):
    oc = numeric_dim(old_p["C"], fallback=old_p["L"] if isinstance(old_p["L"], (int, float)) else 0.0)
    ol = old_p["L"] if isinstance(old_p["L"], (int, float)) else 0.0
    oa = old_p["A"] if isinstance(old_p["A"], (int, float)) else 0.0
    nc = new_v["C"] or 0.0
    nl = new_v["L"] or 0.0
    na = new_v["A"] or 0.0
    return abs(nc - oc) + abs(nl - ol) + abs(na - oa)


# --------------------------------------------------------------------------- #
# correspondencia (regra fixa: nome-chave + menor diferenca dimensional)
# --------------------------------------------------------------------------- #

def find_match(product, old_by_category):
    categoria = product["categoria"]
    if not categoria:
        return None
    candidates_pool = old_by_category.get(categoria, [])
    if not candidates_pool:
        return None

    new_tokens = core_tokens(product["display"])
    if not new_tokens:
        return None
    keyword = new_tokens[0]

    candidates = [p for p in candidates_pool if keyword in core_tokens(p["name"])]
    if not candidates:
        return None

    best = None
    for variant in product["variants"]:
        for old_p in candidates:
            score = dimension_diff(variant, old_p)
            if best is None or score < best[0]:
                best = (score, variant, old_p)
    if best is None:
        return None

    score, variant, old_p = best
    tied_brands = sorted({c["brand"] for c in candidates
                          if c["name"] == old_p["name"] and dimension_diff(variant, c) == score})
    brand_label = "/".join(tied_brands) if tied_brands else old_p["brand"]
    return {"variant": variant, "old": old_p, "brand_label": brand_label}


# --------------------------------------------------------------------------- #
# leitura: fonte legada "Precos antigos - Century e .pv" (fallback)
# --------------------------------------------------------------------------- #

# campos de dimensao reconhecidos no texto da "configuracao" (mesmas chaves de
# CONFIG_FIELDS); qualquer outra chave contendo "ACABAMENTO", ou "OPCAO DE TOPO",
# e tratada como o descritor do tipo de acabamento daquela linha de preco.
_LEGACY_DIM_KEYS = {
    "COMPRIMENTO DA MODULACAO": "C",
    "PROFUNDIDADE DA MODULACAO": "L",
    "ALTURA DA MODULACAO": "A",
}
_LEGACY_FINISH_KEY_RE = re.compile(r"ACABAMENTO|OPCAO DE TOPO")


def load_legacy_products():
    """Le 'Preços antigos - Century e .pv.xlsx' e retorna uma lista de
    'configuracoes' (uma por combinacao unica de modelo+modulacao+dimensoes+
    atributos extras), cada uma agregando todos os acabamentos descritos para
    ela em um dict {nome_do_acabamento: preco}:
        {"display", "modulacao", "C", "L", "A", "finishes": {nome: preco}}
    """
    if not LEGACY_PRICES_FILE.exists():
        return []
    wb = openpyxl.load_workbook(LEGACY_PRICES_FILE, data_only=True)
    ws = wb["Export"]

    groups = {}
    order = []
    current_modelo = None
    current_modulacao = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        modelo, modulacao, config_text, preco = (row + (None,) * 4)[:4]
        if modelo:
            current_modelo = modelo
        if modulacao:
            current_modulacao = modulacao
        if not config_text or not isinstance(config_text, str) or "MODULACAO DO PRODUTO" not in config_text:
            continue
        if not isinstance(preco, (int, float)):
            continue

        cleaned = strip_accents(config_text).upper()
        dims = {"C": None, "L": None, "A": None}
        finish_name = None
        extra = []
        for line in cleaned.splitlines():
            line = line.strip()
            if ":" not in line:
                continue
            key, _, val = line.partition(":")
            key, val = key.strip(), val.strip()
            if key in _LEGACY_DIM_KEYS:
                m = re.search(r"[\d.]+", val)
                if m:
                    dims[_LEGACY_DIM_KEYS[key]] = float(m.group(0))
            elif key == "MODULACAO DO PRODUTO" or key == "FORMATO DA MODULACAO":
                continue
            elif _LEGACY_FINISH_KEY_RE.search(key):
                finish_name = val
            else:
                extra.append(f"{key}={val}")
        if finish_name is None or current_modelo is None:
            continue

        gkey = (current_modelo, current_modulacao, dims["C"], dims["L"], dims["A"], tuple(extra))
        if gkey not in groups:
            display = current_modelo.replace("MODELO ", "", 1).title()
            groups[gkey] = {
                "display": display, "modulacao": current_modulacao,
                "C": dims["C"], "L": dims["L"], "A": dims["A"], "finishes": {},
            }
            order.append(gkey)
        groups[gkey]["finishes"][finish_name] = float(preco)

    return [groups[k] for k in order]


def find_legacy_match(product, legacy_products):
    """Mesma regra de correspondencia (nome-chave + menor diferenca dimensional),
    mas sem filtro de categoria: a planilha legada nao separa por categoria/aba,
    e inclui acessorios (ex.: ESPELHO, MANCEBO) fora das 7 categorias principais."""
    new_tokens = core_tokens(product["display"])
    if not new_tokens:
        return None
    keyword = new_tokens[0]

    candidates = [p for p in legacy_products
                  if keyword in core_tokens(p["display"]) or keyword in core_tokens(p["modulacao"] or "")]
    if not candidates:
        return None

    best = None
    for variant in product["variants"]:
        for old_p in candidates:
            score = dimension_diff(variant, old_p)
            if best is None or score < best[0]:
                best = (score, variant, old_p)
    if best is None:
        return None

    _, variant, old_p = best
    return {"variant": variant, "old": old_p}


# --------------------------------------------------------------------------- #
# montagem dos dados de comparacao
# --------------------------------------------------------------------------- #

def build_rows(variant, old_p, finish_order):
    """Linhas de comparacao contra uma das 3 planilhas de marca (colunas fixas
    de acabamento). Quando o acabamento novo nao tem correspondente preenchido
    na planilha antiga, a linha e listada apenas como informativa (sem %)."""
    rows = []
    for finish in finish_order:
        new_price = variant["finishes"].get(finish)
        if new_price is None:
            continue
        mapping = FINISH_MAP.get(finish) or ESP_FINISH_MAP.get(finish)
        if mapping is None:
            continue
        is_esp = finish in ESP_FINISH_MAP
        col_idx, factor = mapping
        old_raw = old_p["cols"][col_idx - COL_VIDRO_ESPELHO]
        if not isinstance(old_raw, (int, float)):
            rows.append({
                "finish": finish,
                "note": FINISH_NOTES.get(finish, "sem equivalente preenchido na tabela anterior"),
                "old": None, "new": new_price, "delta": None, "pct": None,
                "info_only": True, "is_esp": is_esp,
            })
            continue
        old_price = old_raw * factor
        delta = new_price - old_price
        pct = (delta / old_price) * 100 if old_price else 0.0
        rows.append({
            "finish": finish,
            "note": FINISH_NOTES.get(finish),
            "old": old_price, "new": new_price,
            "delta": delta, "pct": pct,
            "info_only": False, "is_esp": is_esp,
        })
    return rows


def build_legacy_rows(variant, legacy_p, finish_order):
    """Linhas de comparacao contra a fonte legada 'Precos antigos - Century e
    .pv'. Ali o acabamento do tampo vem descrito na propria configuracao (nao em
    colunas fixas), entao a comparacao e feita por NOME (mesmo vocabulario de
    'Novos precos', com pequenas variacoes tratadas em LEGACY_FINISH_ALIASES).

    Caso especial (instrucao do usuario): quando o produto novo nao diferencia
    acabamentos e so tem o preco "NENHUM" preenchido (modelos "agrupados", ex.
    SAMMY), esse preco unico e comparado contra CADA acabamento listado na fonte
    legada para aquela configuracao - gerando uma linha por acabamento antigo.

    Quando um acabamento existe de um lado mas nao do outro, a linha e listada
    apenas como informativa (sem calculo de %), nos dois sentidos."""
    new_finishes = {f: variant["finishes"][f] for f in finish_order if f in variant["finishes"]}
    old_finishes = dict(legacy_p["finishes"])
    if not new_finishes or not old_finishes:
        return []

    rows = []

    if set(new_finishes) == {"NENHUM"}:
        nenhum_price = new_finishes["NENHUM"]
        for old_name, old_price in old_finishes.items():
            delta = nenhum_price - old_price
            pct = (delta / old_price) * 100 if old_price else 0.0
            rows.append({
                "finish": old_name,
                "note": "produto novo não diferencia acabamentos — comparado ao preço único (NENHUM)",
                "old": old_price, "new": nenhum_price, "delta": delta, "pct": pct,
                "info_only": False, "is_esp": False,
            })
        return rows

    matched_old = set()
    for finish in finish_order:
        new_price = new_finishes.get(finish)
        if new_price is None:
            continue
        is_esp = finish in ESP_FINISH_MAP
        norm = normalize_finish_name(finish)
        old_match = None
        for old_name in old_finishes:
            on = normalize_finish_name(old_name)
            if on == norm or LEGACY_FINISH_ALIASES.get(on) == norm or LEGACY_FINISH_ALIASES.get(norm) == on:
                old_match = old_name
                break
        if old_match is not None:
            old_price = old_finishes[old_match]
            matched_old.add(old_match)
            delta = new_price - old_price
            pct = (delta / old_price) * 100 if old_price else 0.0
            rows.append({
                "finish": finish, "note": None,
                "old": old_price, "new": new_price, "delta": delta, "pct": pct,
                "info_only": False, "is_esp": is_esp,
            })
        else:
            rows.append({
                "finish": finish, "note": "sem equivalente na tabela anterior",
                "old": None, "new": new_price, "delta": None, "pct": None,
                "info_only": True, "is_esp": is_esp,
            })

    for old_name, old_price in old_finishes.items():
        if old_name in matched_old:
            continue
        rows.append({
            "finish": old_name, "note": "sem equivalente na tabela atual",
            "old": old_price, "new": None, "delta": None, "pct": None,
            "info_only": True, "is_esp": False,
        })
    return rows


def pick_representative_variant(product):
    """Para produtos 'Sem Correspondente' (sem nenhuma referencia antiga), escolhe
    UMA variante para exibir - mesma logica de 'uma config por produto' usada nos
    produtos comparados. Mesa de Jantar -> variante mais proxima de 2,70m de
    comprimento (tamanho de referencia do segmento); demais -> variante mais
    proxima da dimensao media (C+L+A) do proprio produto."""
    variants = product["variants"]
    if len(variants) == 1:
        return variants[0]
    if product["categoria"] == "Mesa de Jantar":
        with_c = [v for v in variants if isinstance(v["C"], (int, float))]
        if with_c:
            return min(with_c, key=lambda v: abs(v["C"] - 2.70))

    def total_dim(v):
        return (v["C"] or 0.0) + (v["L"] or 0.0) + (v["A"] or 0.0)

    avg = sum(total_dim(v) for v in variants) / len(variants)
    return min(variants, key=lambda v: abs(total_dim(v) - avg))


def build_report_data():
    new_products = load_new_products()
    old_by_category = load_old_products()
    legacy_products = load_legacy_products()

    matched_by_cat = {cat: [] for cat in CATEGORY_ORDER}
    category_order = list(CATEGORY_ORDER)
    unmatched = []

    def ensure_cat(cat):
        if cat not in matched_by_cat:
            matched_by_cat[cat] = []
            category_order.append(cat)
        return matched_by_cat[cat]

    for product in new_products:
        match = find_match(product, old_by_category)
        if match is not None:
            rows = build_rows(match["variant"], match["old"], FINISH_ORDER_ESP)
            if rows:
                ensure_cat(product["categoria"]).append({
                    "display": product["display"],
                    "brand_label": match["brand_label"],
                    "size": fmt_size(match["variant"]["C"], match["variant"]["L"], match["variant"]["A"]),
                    "ref": match["old"]["name"],
                    "rows": rows,
                })
                continue

        legacy_match = find_legacy_match(product, legacy_products)
        if legacy_match is not None:
            rows = build_legacy_rows(legacy_match["variant"], legacy_match["old"], FINISH_ORDER_ESP)
            if rows:
                cat = product["categoria"] or category_label(product["categoria_raw"])
                ref_parts = [legacy_match["old"]["display"]]
                if legacy_match["old"]["modulacao"]:
                    ref_parts.append(legacy_match["old"]["modulacao"].title())
                ensure_cat(cat).append({
                    "display": product["display"],
                    "brand_label": "Century/Pv (legado)",
                    "size": fmt_size(legacy_match["variant"]["C"], legacy_match["variant"]["L"], legacy_match["variant"]["A"]),
                    "ref": " · ".join(ref_parts),
                    "rows": rows,
                })
                continue

        unmatched.append(product)

    return matched_by_cat, unmatched, category_order


# --------------------------------------------------------------------------- #
# geracao do HTML
# --------------------------------------------------------------------------- #

PAGE_CSS = """
:root{--dg:#484c40;--mg:#84867b;--kh:#878264;--sd:#c5c0b3;--ow:#dcd8d3;--cr:#f2f0ec;--bg:#ede9e4;--ch:#211f1e;--wm:#494038;--bl:#909d9c;--up:#2e6b35;--upbg:#ebf4ec;--upbd:#b8dbbf;--dn:#8b2020;--dnbg:#fdf0f0;--dnbd:#e8b4b4;--nt:#5a5849;--ntbg:#f5f3ef;--ntbd:#ccc9c0;--ep:#4a3a6b;--epbg:#f0ecfa;--epbd:#c4b4e8;--nm:#5a4520;--nmbg:#faf4eb;--nmbd:#d4b87a}
*{margin:0;padding:0;box-sizing:border-box}html{scroll-behavior:smooth}
body{font-family:'Albert Sans',sans-serif;background:var(--cr);color:var(--ch);font-size:13px;line-height:1.6}
.hdr{background:var(--ch);color:#fff;padding:52px 64px 44px;position:relative;overflow:hidden}
.hdr::after{content:'';position:absolute;top:-120px;right:-120px;width:400px;height:400px;border-radius:50%;background:rgba(255,255,255,.03);pointer-events:none}
.hdr-eye{font-size:9px;font-weight:600;letter-spacing:.3em;text-transform:uppercase;color:var(--sd);margin-bottom:14px;opacity:.8}
.hdr-title{font-family:'Geologica',sans-serif;font-size:42px;font-weight:200;letter-spacing:-.02em;line-height:1.05;margin-bottom:6px}
.hdr-title b{font-weight:700}.hdr-sub{color:var(--sd);font-size:13px;font-weight:300;opacity:.8;margin-bottom:32px}
.hdr-pills{display:flex;flex-wrap:wrap;gap:28px;border-top:1px solid rgba(255,255,255,.1);padding-top:24px}
.pill{display:flex;flex-direction:column;gap:3px}.pill-l{font-size:8px;letter-spacing:.22em;text-transform:uppercase;color:rgba(255,255,255,.4)}.pill-v{font-size:13px;font-weight:500;color:rgba(255,255,255,.85)}
.hdr-brand{position:absolute;right:64px;bottom:44px;font-family:'Geologica',sans-serif;font-size:13px;font-weight:200;letter-spacing:.4em;text-transform:uppercase;color:rgba(255,255,255,.18)}.hdr-brand b{font-weight:700;color:rgba(255,255,255,.32)}
.nav{position:sticky;top:0;z-index:200;background:var(--dg);display:flex;overflow-x:auto;scrollbar-width:none;border-bottom:1px solid rgba(0,0,0,.2)}.nav::-webkit-scrollbar{display:none}
.nav-home{padding:13px 18px 13px 24px;font-family:'Geologica',sans-serif;font-size:10px;font-weight:700;letter-spacing:.25em;text-transform:uppercase;color:rgba(255,255,255,.55);text-decoration:none;white-space:nowrap;border-right:1px solid rgba(255,255,255,.1);margin-right:8px}
.nav-a{padding:13px 16px;font-size:10px;font-weight:600;letter-spacing:.1em;text-transform:uppercase;color:rgba(255,255,255,.45);text-decoration:none;white-space:nowrap;border-bottom:2px solid transparent;transition:all .2s}.nav-a:hover{color:#fff;border-bottom-color:var(--sd)}
.nav-a.nm-nav{color:rgba(212,184,122,.6)}.nav-a.nm-nav:hover{color:#d4b87a;border-bottom-color:#d4b87a}
.main{max-width:1180px;margin:0 auto;padding:52px 64px 96px}
.eyebrow{font-family:'Geologica',sans-serif;font-size:9px;font-weight:600;letter-spacing:.28em;text-transform:uppercase;color:var(--kh);margin-bottom:18px}
.note{background:#fff;border:1px solid var(--ow);border-left:3px solid var(--kh);padding:14px 18px;margin-bottom:36px;font-size:12px;color:var(--wm);line-height:1.8}.note b{font-weight:600;color:var(--dg)}
.cgrid{display:grid;grid-template-columns:repeat(auto-fill,minmax(172px,1fr));gap:10px;margin-bottom:60px}
.ccard{background:#fff;border:1px solid var(--ow);padding:20px 18px 16px;cursor:pointer;transition:border-color .2s,box-shadow .2s}.ccard:hover{border-color:var(--sd);box-shadow:0 2px 12px rgba(0,0,0,.06)}
.ccard-nm{font-family:'Geologica',sans-serif;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.1em;color:var(--dg);margin-bottom:1px}.ccard-sub{font-size:10px;color:var(--mg);margin-bottom:14px}
.ccard-avg{font-family:'Geologica',sans-serif;font-size:28px;font-weight:200;line-height:1;margin-bottom:2px}.ccard-avg.up{color:var(--up)}.ccard-avg.down{color:var(--dn)}.ccard-avg.neutral{color:var(--nt)}
.ccard-lbl{font-size:9px;letter-spacing:.12em;text-transform:uppercase;color:var(--mg);margin-bottom:12px}
.ccard-rng{display:flex;flex-wrap:wrap;align-items:flex-start;gap:4px;font-size:10px}.ccard-rng span{display:flex;flex-direction:column;gap:1px}.ccard-rng em{font-style:normal;font-size:9px;color:var(--mg)}.ccard-rng .up{color:var(--up)}.ccard-rng .down{color:var(--dn)}.ccard-rng .neutral{color:var(--nt)}.sep{color:var(--ow);padding:0 3px;align-self:center}
.csec{margin-bottom:60px}.csec-hdr{display:flex;align-items:center;gap:14px;padding-bottom:14px;margin-bottom:20px;border-bottom:2px solid var(--ow)}.csec-hdr h2{font-family:'Geologica',sans-serif;font-size:22px;font-weight:300;color:var(--ch);letter-spacing:-.01em}
.cbadge{font-size:11px;font-weight:600;padding:3px 11px;border-radius:20px}.cbadge.up{background:var(--upbg);color:var(--up);border:1px solid var(--upbd)}.cbadge.down{background:var(--dnbg);color:var(--dn);border:1px solid var(--dnbd)}.cbadge.neutral{background:var(--ntbg);color:var(--nt);border:1px solid var(--ntbd)}
.nm-intro{font-size:12px;color:var(--nm);background:var(--nmbg);border:1px solid var(--nmbd);padding:10px 16px;margin-bottom:16px;border-radius:2px}
.pblock{background:#fff;border:1px solid var(--ow);margin-bottom:12px;overflow:hidden}
.ph{padding:12px 18px 10px;background:var(--bg);border-bottom:1px solid var(--ow)}
.pname{font-family:'Geologica',sans-serif;font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:var(--ch);margin-bottom:5px}
.pmeta{display:flex;flex-wrap:wrap;gap:5px;align-items:center}.tag{font-size:9px;font-weight:600;padding:2px 8px;letter-spacing:.05em}.tag.brand{background:var(--ch);color:rgba(255,255,255,.6)}.tag.size{background:var(--dg);color:rgba(255,255,255,.7)}.tag.ref{font-weight:400;font-style:italic;background:transparent;border:1px solid var(--ow);color:var(--mg)}
.ft{width:100%;border-collapse:collapse}.ft thead th{background:var(--ch);color:rgba(255,255,255,.5);font-size:9px;font-weight:600;letter-spacing:.18em;text-transform:uppercase;padding:8px 14px;text-align:left}.th-pr,.th-vr,.th-ab{text-align:right!important}
.ft tbody tr{border-bottom:1px solid var(--ow)}.ft tbody tr:last-child{border:none}.ft tbody tr:hover{background:var(--cr)}.ft td{padding:9px 14px;vertical-align:middle}
.fn{font-size:12px;color:var(--ch);min-width:160px}.fn-note{font-size:9px;color:var(--bl);font-style:italic;margin-left:6px}
.pr{text-align:right;font-variant-numeric:tabular-nums;font-size:12px}.pr.old{color:var(--mg)}.pr.new{color:var(--ch);font-weight:500}
.vr{text-align:right;font-size:12px;font-weight:700;min-width:80px}.ab{text-align:right;font-size:11px;min-width:100px}
.up{color:var(--up)}.down{color:var(--dn)}.neutral{color:var(--nt)}
.pblock.no-match{border-color:var(--nmbd);border-left:3px solid var(--nmbd)}.pblock.no-match .ph{background:var(--nmbg)}.nm-badge{font-size:9px;font-weight:500;background:var(--nmbg);border:1px solid var(--nmbd);color:var(--nm);padding:2px 8px;border-radius:3px;letter-spacing:.04em;vertical-align:middle;margin-left:8px}
.ftr{background:var(--ch);color:rgba(255,255,255,.3);padding:24px 64px;display:flex;justify-content:space-between;align-items:center;font-size:11px}.ftr-brand{font-family:'Geologica',sans-serif;letter-spacing:.3em;text-transform:uppercase;font-weight:200;color:rgba(255,255,255,.2)}.ftr-brand b{font-weight:700}
.dual-esp{display:contents}.dual-noesp{display:none}
body.no-esp .dual-esp{display:none}body.no-esp .dual-noesp{display:contents}
body.no-esp tr.esp-row{display:none}body.no-esp .pblock.esp-only{display:none}
.esp-switch{display:inline-flex;align-items:center;gap:10px;background:#fff;border:1px solid var(--ow);padding:8px 18px 8px 10px;margin-bottom:18px;cursor:pointer;font:inherit;color:var(--dg);border-radius:30px;transition:border-color .2s}.esp-switch:hover{border-color:var(--sd)}
.esp-switch-track{width:36px;height:20px;border-radius:12px;background:var(--ow);position:relative;transition:background .2s;flex:none}
.esp-switch-thumb{position:absolute;top:2px;left:2px;width:16px;height:16px;border-radius:50%;background:#fff;box-shadow:0 1px 3px rgba(0,0,0,.25);transition:left .2s}
.esp-switch[aria-checked="true"] .esp-switch-track{background:var(--ep)}.esp-switch[aria-checked="true"] .esp-switch-thumb{left:18px}
.esp-switch-label{font-size:11px;font-weight:600;letter-spacing:.03em}
tr.info-only .vr,tr.info-only .ab{color:var(--bl)}
"""


def dual(esp_html, noesp_html, tag="span"):
    """Renderiza dois trechos de HTML lado a lado, um visivel com Marmores
    Especiais (ESP) inclusos e outro sem - o interruptor no topo da pagina
    alterna a classe 'no-esp' no <body> e o CSS troca qual trecho fica visivel.
    Usa display:contents para nao afetar o layout do conteudo interno."""
    return f'<{tag} class="dual-esp">{esp_html}</{tag}><{tag} class="dual-noesp">{noesp_html}</{tag}>'


def has_non_esp_rows(product):
    return any(not r["is_esp"] for r in product["rows"])


def category_stats(products, include_esp):
    """Media/melhor/pior variacao percentual e contagem de produtos, considerando
    (ou nao) as linhas de Marmore Especial - usado para alimentar as duas vistas
    (com/sem ESP) dos cards de resumo e badges de categoria."""
    rows = []
    count = 0
    for p in products:
        prows = [r for r in p["rows"] if r["pct"] is not None and (include_esp or not r["is_esp"])]
        if prows:
            count += 1
            rows.extend((r, p["display"]) for r in prows)
    if not rows:
        return None
    pcts = [r["pct"] for r, _ in rows]
    return {
        "avg": sum(pcts) / len(pcts),
        "worst": min(rows, key=lambda x: x[0]["pct"]),
        "best": max(rows, key=lambda x: x[0]["pct"]),
        "count": count,
    }


def render_summary_cards(matched_by_cat, category_order):
    cards = []
    for cat in category_order:
        products = matched_by_cat.get(cat) or []
        if not products:
            continue
        stats_all = category_stats(products, True)
        stats_noesp = category_stats(products, False)
        if stats_all is None and stats_noesp is None:
            continue

        def card_body(stats):
            if stats is None:
                return ('<div class="ccard-sub">0 produtos</div>'
                        '<div class="ccard-avg neutral">—</div>'
                        '<div class="ccard-lbl">sem variação calculável</div>')
            unit = "produto" if stats["count"] == 1 else "produtos"
            return (f'<div class="ccard-sub">{stats["count"]} {unit}</div>'
                    f'<div class="ccard-avg {variation_class(stats["avg"])}">{fmt_pct(stats["avg"])}</div>'
                    f'<div class="ccard-lbl">variação média</div>'
                    f'<div class="ccard-rng">'
                    f'<span class="{variation_class(stats["worst"][0]["pct"])}">{fmt_pct(stats["worst"][0]["pct"])} <em>{short_name(stats["worst"][1])}</em></span>'
                    f'<span class="sep">·</span>'
                    f'<span class="{variation_class(stats["best"][0]["pct"])}">{fmt_pct(stats["best"][0]["pct"])} <em>{short_name(stats["best"][1])}</em></span>'
                    f'</div>')

        cards.append(f"""<div class="ccard" onclick="go('{cat_slug(cat)}')">
  <div class="ccard-nm">{cat}</div>
  {dual(card_body(stats_all), card_body(stats_noesp), tag="div")}
</div>""")
    return "".join(cards)


def render_product_block(product):
    rows_html = []
    for r in product["rows"]:
        classes = []
        if r["is_esp"]:
            classes.append("esp-row")
        if r["info_only"]:
            classes.append("info-only")
        cls_attr = f' class="{" ".join(classes)}"' if classes else ""
        note = f'<span class="fn-note">{r["note"]}</span>' if r["note"] else ""
        if r["info_only"]:
            old_cell = fmt_brl(r["old"]) if r["old"] is not None else "—"
            new_cell = fmt_brl(r["new"]) if r["new"] is not None else "—"
            rows_html.append(f"""<tr{cls_attr}>
  <td class="fn">{r['finish']}{note}</td>
  <td class="pr old">{old_cell}</td><td class="pr new">{new_cell}</td>
  <td class="vr neutral">—</td><td class="ab neutral">—</td></tr>""")
        else:
            rows_html.append(f"""<tr{cls_attr}>
  <td class="fn">{r['finish']}{note}</td>
  <td class="pr old">{fmt_brl(r['old'])}</td><td class="pr new">{fmt_brl(r['new'])}</td>
  <td class="vr {variation_class(r['pct'])}">{arrow(r['pct'])} {fmt_pct(r['pct'])}</td>
  <td class="ab {'up' if r['delta'] >= 0 else 'down'}">{fmt_brl_signed(r['delta'])}</td></tr>""")
    esp_only_cls = " esp-only" if not has_non_esp_rows(product) else ""
    return f"""<div class="pblock{esp_only_cls}">
  <div class="ph"><div class="pname">{product['display']}</div>
  <div class="pmeta"><span class="tag brand">{product['brand_label']}</span><span class="tag size">{product['size']}</span><span class="tag ref">{product['ref']}</span></div></div>
  <table class="ft"><thead><tr>
    <th class="th-fn">Acabamento</th>
    <th class="th-pr">Preço Anterior</th><th class="th-pr">Preço Novo</th>
    <th class="th-vr">Var %</th><th class="th-ab">Var R$</th>
  </tr></thead><tbody>{''.join(rows_html)}</tbody></table>
</div>"""


def render_category_section(cat, products):
    if not products:
        return ""
    stats_all = category_stats(products, True)
    stats_noesp = category_stats(products, False)
    slug = cat_slug(cat)
    blocks = "".join(render_product_block(p) for p in products)

    def badge(stats):
        if stats is None:
            return '<span class="cbadge neutral">sem variação calculável</span>'
        return f'<span class="cbadge {variation_class(stats["avg"])}">{fmt_pct(stats["avg"])} média</span>'

    return f"""<section class="csec" id="sec-{slug}"><div class="csec-hdr"><h2>{cat}</h2>{dual(badge(stats_all), badge(stats_noesp))}</div>{blocks}</section>"""


def render_unmatched(unmatched):
    if not unmatched:
        return ""
    blocks = []
    for product in unmatched:
        variant = pick_representative_variant(product)
        rows = []
        for finish in FINISH_ORDER_ESP:
            price = variant["finishes"].get(finish)
            if price is None:
                continue
            cls_attr = ' class="esp-row"' if finish in ESP_FINISH_MAP else ""
            rows.append(f'<tr{cls_attr}><td class="fn">{finish}</td><td class="pr new" colspan="4">{fmt_brl(price)}</td></tr>')
        if not rows:
            continue
        size = fmt_size(variant["C"], variant["L"], variant["A"])
        blocks.append(f"""<div class="pblock no-match">
  <div class="ph"><div class="pname">{product['display']} <span class="nm-badge">Sem correspondente — apenas listagem</span></div>
  <div class="pmeta"><span class="tag brand">—</span><span class="tag size">{size}</span></div></div>
  <table class="ft"><thead><tr><th class="th-fn">Acabamento</th><th class="th-pr" colspan="4">Preço Novo (configuração analisada)</th></tr></thead><tbody>{''.join(rows)}</tbody></table>
</div>""")
    return f"""<section class="csec" id="sec-sem-correspondente">
  <div class="csec-hdr"><h2>Sem Correspondente</h2><span class="cbadge neutral">apenas listagem</span></div>
  <div class="nm-intro">Produtos abaixo não possuem equivalente em nenhuma das fontes antigas (3 planilhas de marca + fonte legada Century/PV). É exibida apenas a configuração de tamanho mais representativa do produto (tamanho médio do produto, ou variante mais próxima de 2,70m para mesas de jantar), com os preços novos para referência.</div>
  {''.join(blocks)}</section>"""


def render_page(matched_by_cat, unmatched, category_order):
    total_matched_all = sum(len(v) for v in matched_by_cat.values())
    total_matched_noesp = sum(1 for v in matched_by_cat.values() for p in v if has_non_esp_rows(p))
    total_unmatched = len(unmatched)

    nav_links = "".join(
        f'<a class="nav-a" href="#sec-{cat_slug(cat)}">{cat}</a>'
        for cat in category_order if matched_by_cat.get(cat)
    )

    note = """<div class="note">
    <b>Correções aplicadas:</b>&nbsp;
    <b>Mármore Fornecido</b> → preço sem topo −8%.&nbsp;
    <b>Vidro Normal</b> → mesmo preço do sem topo (coluna "Tampo: Laca ou Lamina com ou sem Vidro Normal").&nbsp;
    <b>Fonte legada (Century/PV)</b> → consultada apenas quando não há correspondência nas 3 planilhas de marca; comparação direta por nome de acabamento — quando o produto novo não diferencia acabamentos (apenas "Nenhum" preenchido), esse preço único é comparado a cada acabamento disponível na tabela antiga.&nbsp;
    <b>Sem equivalente</b> → quando um acabamento existe de um lado mas não do outro, a linha é listada apenas como informação, sem cálculo de variação.&nbsp;
    <b>Sem correspondente</b> → agrupados ao final, exibindo a configuração de tamanho mais representativa de cada produto.&nbsp;
    Use o interruptor acima para incluir ou excluir os <b>Mármores Especiais (ESP)</b> da análise e dos cálculos de variação — os valores se recalculam automaticamente.
  </div>"""

    switch = """<button id="esp-switch" class="esp-switch" type="button" role="switch" aria-checked="true" onclick="toggleEsp()">
    <span class="esp-switch-track"><span class="esp-switch-thumb"></span></span>
    <span class="esp-switch-label">Incluir Mármores Especiais (ESP) na análise</span>
  </button>"""

    sections = "".join(render_category_section(cat, matched_by_cat[cat]) for cat in category_order)

    pill_comparativo = dual(
        f'<span class="pill-v">{total_matched_all} produtos</span>',
        f'<span class="pill-v">{total_matched_noesp} produtos</span>',
    )

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Análise Comparativa de Preços — Wood | SoHome</title>
<link href="https://fonts.googleapis.com/css2?family=Geologica:wght@200;300;400;600;700&family=Albert+Sans:ital,wght@0,300;0,400;0,500;0,600;1,300;1,400&display=swap" rel="stylesheet">
<style>{PAGE_CSS}</style>
</head>
<body>
<header class="hdr">
  <div class="hdr-eye">Grupo SoHome · Linha Wood</div>
  <h1 class="hdr-title">Análise Comparativa<br><b>de Preços</b></h1>
  <p class="hdr-sub">Reprecificação vs. Tabelas 28/04 — por categoria e acabamento</p>
  <div class="hdr-pills">
    <div class="pill"><span class="pill-l">Com comparativo</span>{pill_comparativo}</div>
    <div class="pill"><span class="pill-l">Sem correspondente</span><span class="pill-v">{total_unmatched} produtos</span></div>
    <div class="pill"><span class="pill-l">Fontes antigas</span><span class="pill-v">Century · PV · Private Label 28/04 · Legado Century/PV</span></div>
  </div>
  <div class="hdr-brand">SO<b>HOME</b></div>
</header>
<nav class="nav">
  <a class="nav-home" href="#top">SO<b>HOME</b></a>
  <a class="nav-a" href="#resumo">Resumo</a>{nav_links}<a class="nav-a nm-nav" href="#sec-sem-correspondente">Sem Correspondente</a>
</nav>
<main class="main" id="top">
<section id="resumo" style="margin-bottom:60px">
  <p class="eyebrow">Resumo por Categoria</p>
  {switch}
  {note}
  <div class="cgrid">{render_summary_cards(matched_by_cat, category_order)}</div>
</section>
<section><p class="eyebrow">Detalhamento por Produto</p>{sections}{render_unmatched(unmatched)}</section>
</main>
<footer class="ftr">
  <span class="ftr-brand">SO<b>HOME</b> · Wood · Análise de Preços</span>
  <span>Comparativo: Reprecificação vs. Tabelas 28/04</span>
</footer>
<script>
function go(slug){{const el=document.getElementById('sec-'+slug);if(el)el.scrollIntoView({{behavior:'smooth',block:'start'}});}}
function toggleEsp(){{
  const noEsp = document.body.classList.toggle('no-esp');
  document.getElementById('esp-switch').setAttribute('aria-checked', noEsp ? 'false' : 'true');
}}
</script>
</body></html>
"""


# --------------------------------------------------------------------------- #
# main
# --------------------------------------------------------------------------- #

def main():
    print("Lendo planilhas e calculando comparações (com e sem Mármore ESP via interruptor)...")
    matched, unmatched, category_order = build_report_data()
    OUTPUT_REPORT.write_text(render_page(matched, unmatched, category_order), encoding="utf-8")
    total_matched = sum(len(v) for v in matched.values())
    print(f"  -> {OUTPUT_REPORT.name}  "
          f"({total_matched} com comparativo, {len(unmatched)} sem correspondente)")


if __name__ == "__main__":
    main()
